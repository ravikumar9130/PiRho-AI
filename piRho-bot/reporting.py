"""
Trade Logging and Reporting for Crypto Trading Bot
Handles trade logs, daily reports, and performance tracking
"""

import logging
import os
import datetime
import pandas as pd
from typing import Optional, Dict, Any, List

try:
    from openpyxl import Workbook, load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = logging.getLogger(__name__)

# Default paths
TRADE_LOG_PATH = "trade_log.xlsx"
PERFORMANCE_LOG_PATH = "performance_log.csv"


def initialize_trade_log(path: str = TRADE_LOG_PATH):
    """
    Initialize the trade log Excel file if it doesn't exist.
    
    Args:
        path: Path to the Excel file
    """
    if not HAS_OPENPYXL:
        logger.warning("openpyxl not installed - trade logging to Excel disabled")
        return
    
    if os.path.exists(path):
        logger.debug(f"Trade log already exists at {path}")
        return
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"
        
        # Headers
        headers = [
            "Timestamp", "OrderID", "Symbol", "TradeType", "Strategy",
            "EntryPrice", "ExitPrice", "Quantity", "Leverage",
            "ProfitLoss", "ProfitLossPercent", "ExitReason",
            "Status", "PaperTrade", "Rationale"
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Summary sheet
        ws_summary = wb.create_sheet("Summary")
        summary_headers = [
            "Date", "TotalTrades", "WinningTrades", "LosingTrades",
            "WinRate", "TotalPnL", "AveragePnL", "BestTrade", "WorstTrade"
        ]
        for col, header in enumerate(summary_headers, 1):
            ws_summary.cell(row=1, column=col, value=header)
        
        wb.save(path)
        logger.info(f"Trade log initialized at {path}")
        
    except Exception as e:
        logger.error(f"Failed to initialize trade log: {e}")


def log_trade(trade_details: Dict[str, Any], path: str = TRADE_LOG_PATH):
    """
    Log a completed trade to the Excel file.
    
    Args:
        trade_details: Dictionary with trade information
        path: Path to the Excel file
    """
    if not HAS_OPENPYXL:
        # Fallback to CSV logging
        _log_trade_csv(trade_details)
        return
    
    try:
        if not os.path.exists(path):
            initialize_trade_log(path)
        
        wb = load_workbook(path)
        ws = wb["Trades"]
        
        # Find next empty row
        next_row = ws.max_row + 1
        
        # Write trade data
        columns = [
            "Timestamp", "OrderID", "Symbol", "TradeType", "Strategy",
            "EntryPrice", "ExitPrice", "Quantity", "Leverage",
            "ProfitLoss", "ProfitLossPercent", "ExitReason",
            "Status", "PaperTrade", "Rationale"
        ]
        
        for col, key in enumerate(columns, 1):
            value = trade_details.get(key, "")
            ws.cell(row=next_row, column=col, value=value)
        
        wb.save(path)
        logger.info(f"Trade logged: {trade_details.get('Symbol')} P&L: ${trade_details.get('ProfitLoss', 0):.2f}")
        
    except Exception as e:
        logger.error(f"Failed to log trade: {e}")
        _log_trade_csv(trade_details)


def _log_trade_csv(trade_details: Dict[str, Any], path: str = "trades.csv"):
    """Fallback CSV logging."""
    try:
        import csv
        
        file_exists = os.path.exists(path)
        
        with open(path, 'a', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=trade_details.keys())
            
            if not file_exists:
                writer.writeheader()
            
            writer.writerow(trade_details)
        
        logger.info(f"Trade logged to CSV: {path}")
        
    except Exception as e:
        logger.error(f"Failed to log trade to CSV: {e}")


def get_trade_history(
    path: str = TRADE_LOG_PATH,
    days: int = 30
) -> pd.DataFrame:
    """
    Get trade history from the log.
    
    Args:
        path: Path to the trade log
        days: Number of days to look back
        
    Returns:
        DataFrame with trade history
    """
    try:
        if not os.path.exists(path):
            return pd.DataFrame()
        
        if HAS_OPENPYXL:
            df = pd.read_excel(path, sheet_name="Trades")
        else:
            csv_path = path.replace('.xlsx', '.csv')
            if os.path.exists(csv_path):
                df = pd.read_csv(csv_path)
            else:
                return pd.DataFrame()
        
        if df.empty:
            return df
        
        # Filter by date
        df['Timestamp'] = pd.to_datetime(df['Timestamp'])
        cutoff = datetime.datetime.now() - datetime.timedelta(days=days)
        df = df[df['Timestamp'] >= cutoff]
        
        return df
        
    except Exception as e:
        logger.error(f"Failed to get trade history: {e}")
        return pd.DataFrame()


def calculate_performance_metrics(trades_df: pd.DataFrame) -> Dict[str, Any]:
    """
    Calculate performance metrics from trade history.
    
    Args:
        trades_df: DataFrame with trade history
        
    Returns:
        Dictionary with performance metrics
    """
    if trades_df.empty:
        return {
            'total_trades': 0,
            'winning_trades': 0,
            'losing_trades': 0,
            'win_rate': 0,
            'total_pnl': 0,
            'average_pnl': 0,
            'best_trade': 0,
            'worst_trade': 0,
            'profit_factor': 0,
            'max_drawdown': 0,
        }
    
    pnl = trades_df['ProfitLoss'].fillna(0)
    
    winning = pnl > 0
    losing = pnl < 0
    
    total_wins = pnl[winning].sum()
    total_losses = abs(pnl[losing].sum())
    
    # Calculate drawdown
    cumulative = pnl.cumsum()
    running_max = cumulative.cummax()
    drawdown = (running_max - cumulative).max()
    
    return {
        'total_trades': len(trades_df),
        'winning_trades': winning.sum(),
        'losing_trades': losing.sum(),
        'win_rate': (winning.sum() / len(trades_df) * 100) if len(trades_df) > 0 else 0,
        'total_pnl': pnl.sum(),
        'average_pnl': pnl.mean(),
        'best_trade': pnl.max(),
        'worst_trade': pnl.min(),
        'profit_factor': (total_wins / total_losses) if total_losses > 0 else float('inf'),
        'max_drawdown': drawdown,
    }


def generate_daily_summary(date: datetime.date = None) -> Dict[str, Any]:
    """
    Generate a daily trading summary.
    
    Args:
        date: Date to summarize (default: today)
        
    Returns:
        Dictionary with daily summary
    """
    if date is None:
        date = datetime.date.today()
    
    trades_df = get_trade_history(days=1)
    
    if trades_df.empty:
        return {
            'date': date.isoformat(),
            'total_trades': 0,
            'net_pnl': 0,
            'message': 'No trades today'
        }
    
    # Filter to specific date
    trades_df['date'] = pd.to_datetime(trades_df['Timestamp']).dt.date
    daily_trades = trades_df[trades_df['date'] == date]
    
    metrics = calculate_performance_metrics(daily_trades)
    
    return {
        'date': date.isoformat(),
        **metrics,
        'strategies_used': daily_trades['Strategy'].unique().tolist() if 'Strategy' in daily_trades else [],
        'symbols_traded': daily_trades['Symbol'].unique().tolist() if 'Symbol' in daily_trades else [],
    }


def send_daily_report(
    config: dict,
    date: str = None,
    starting_capital: float = 0,
    no_trades_reason: str = None
):
    """
    Send daily trading report.
    
    Args:
        config: Configuration dictionary
        date: Date string
        starting_capital: Starting capital for the day
        no_trades_reason: Reason if no trades were made
    """
    try:
        if date is None:
            date = datetime.date.today().isoformat()
        
        summary = generate_daily_summary(datetime.date.fromisoformat(date))
        
        report_lines = [
            f"📊 Daily Trading Report - {date}",
            "=" * 40,
            "",
        ]
        
        if no_trades_reason:
            report_lines.append(f"⚠️ No trades today: {no_trades_reason}")
        else:
            report_lines.extend([
                f"📈 Total Trades: {summary['total_trades']}",
                f"✅ Winning: {summary.get('winning_trades', 0)}",
                f"❌ Losing: {summary.get('losing_trades', 0)}",
                f"🎯 Win Rate: {summary.get('win_rate', 0):.1f}%",
                f"💰 Net P&L: ${summary.get('total_pnl', 0):.2f}",
                f"📊 Best Trade: ${summary.get('best_trade', 0):.2f}",
                f"📉 Worst Trade: ${summary.get('worst_trade', 0):.2f}",
            ])
            
            if starting_capital > 0:
                roi = (summary.get('total_pnl', 0) / starting_capital) * 100
                report_lines.append(f"📈 Daily ROI: {roi:+.2f}%")
            
            if summary.get('strategies_used'):
                report_lines.append(f"🎯 Strategies: {', '.join(summary['strategies_used'])}")
            
            if summary.get('symbols_traded'):
                report_lines.append(f"💹 Symbols: {', '.join(summary['symbols_traded'])}")
        
        report = "\n".join(report_lines)
        logger.info(f"\n{report}")
        
        # TODO: Send via Telegram if configured
        
    except Exception as e:
        logger.error(f"Failed to send daily report: {e}")


def send_monthly_report(config: dict, date: str = None):
    """
    Send monthly trading report.
    
    Args:
        config: Configuration dictionary
        date: Date string (any date in the month)
    """
    try:
        if date is None:
            date = datetime.date.today().isoformat()
        
        target_date = datetime.date.fromisoformat(date)
        
        # Get all trades for the month
        trades_df = get_trade_history(days=31)
        
        if trades_df.empty:
            logger.info("No trades this month for monthly report")
            return
        
        # Filter to current month
        trades_df['month'] = pd.to_datetime(trades_df['Timestamp']).dt.month
        monthly_trades = trades_df[trades_df['month'] == target_date.month]
        
        metrics = calculate_performance_metrics(monthly_trades)
        
        report_lines = [
            f"📅 Monthly Trading Report - {target_date.strftime('%B %Y')}",
            "=" * 50,
            "",
            f"📈 Total Trades: {metrics['total_trades']}",
            f"✅ Winning: {metrics['winning_trades']}",
            f"❌ Losing: {metrics['losing_trades']}",
            f"🎯 Win Rate: {metrics['win_rate']:.1f}%",
            f"💰 Net P&L: ${metrics['total_pnl']:.2f}",
            f"📊 Average P&L: ${metrics['average_pnl']:.2f}",
            f"🏆 Best Trade: ${metrics['best_trade']:.2f}",
            f"📉 Worst Trade: ${metrics['worst_trade']:.2f}",
            f"📈 Profit Factor: {metrics['profit_factor']:.2f}",
            f"📉 Max Drawdown: ${metrics['max_drawdown']:.2f}",
        ]
        
        report = "\n".join(report_lines)
        logger.info(f"\n{report}")
        
    except Exception as e:
        logger.error(f"Failed to send monthly report: {e}")


def get_strategy_performance(days: int = 30) -> Dict[str, Dict[str, Any]]:
    """
    Get performance breakdown by strategy.
    
    Args:
        days: Number of days to analyze
        
    Returns:
        Dictionary with strategy-wise performance
    """
    trades_df = get_trade_history(days=days)
    
    if trades_df.empty or 'Strategy' not in trades_df.columns:
        return {}
    
    result = {}
    
    for strategy in trades_df['Strategy'].unique():
        strategy_trades = trades_df[trades_df['Strategy'] == strategy]
        metrics = calculate_performance_metrics(strategy_trades)
        result[strategy] = metrics
    
    return result
